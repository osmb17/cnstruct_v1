"""
Feature B: Coverage Validation Heatmap.

Writes a visual bar-spacing grid to the Validation tab using openpyxl
conditional formatting. No LLM involved — pure deterministic geometry.

Layout in Validation tab (starting row 4):
  - One grid per bar mark
  - Columns = wall length divisions, rows = wall height divisions
  - Cell colour = green (bar present) / red (gap too large) / white (empty zone)

Works both via xlwings (live) and via openpyxl (file write).
"""

from __future__ import annotations

import math
from dataclasses import dataclass
from typing import Any

from vistadetail.engine.schema import BarRow, Params


# RGB colours
_GREEN  = (198, 239, 206)   # bar present, spacing OK
_YELLOW = (255, 235, 156)   # spacing near max
_RED    = (255, 199, 206)   # spacing exceeds ACI max
_WHITE  = (255, 255, 255)   # no bar zone
_GREY   = (242, 242, 242)   # cover zone (no bar expected)
_BLUE   = (189, 215, 238)   # bar mark header


@dataclass
class HeatmapGrid:
    """One spacing grid for a single bar mark."""
    mark: str
    size: str
    spacing_in: float
    dimension_in: float       # the dimension being divided (height or length)
    cover_in: float
    max_spacing_in: float     # ACI max for this wall thickness
    direction: str            # "horizontal" or "vertical"

    @property
    def bar_positions(self) -> list[float]:
        """Positions of bars along the dimension (inches from face)."""
        usable = self.dimension_in - 2 * self.cover_in
        qty = math.floor(usable / self.spacing_in) + 1
        start = self.cover_in
        return [start + i * self.spacing_in for i in range(qty)]

    @property
    def status(self) -> str:
        if self.spacing_in > self.max_spacing_in:
            return "EXCEEDS_MAX"
        elif self.spacing_in > self.max_spacing_in * 0.85:
            return "NEAR_MAX"
        return "OK"

    @property
    def status_colour(self) -> tuple[int, int, int]:
        return {"OK": _GREEN, "NEAR_MAX": _YELLOW, "EXCEEDS_MAX": _RED}[self.status]


def build_heatmap_grids(bars: list[BarRow], params: Params) -> list[HeatmapGrid]:
    """
    Build HeatmapGrid objects from the generated bar list and validated params.
    Only processes wall-type bars (horizontal and vertical EF marks).
    """
    grids: list[HeatmapGrid] = []
    wall_thick = getattr(params, "wall_thick_in", 9)
    cover      = getattr(params, "cover_in", 2.0)
    max_sp     = min(3 * wall_thick, 18.0)   # ACI 318-19 §24.3.2

    for bar in bars:
        if bar.mark.endswith("H1") or "Horiz" in bar.notes:
            h_sp = getattr(params, "horiz_spacing_in", None)
            h_dim = getattr(params, "wall_height_ft", None)
            if h_sp and h_dim:
                grids.append(HeatmapGrid(
                    mark=bar.mark, size=bar.size,
                    spacing_in=h_sp, dimension_in=h_dim * 12,
                    cover_in=cover, max_spacing_in=max_sp,
                    direction="horizontal",
                ))
        elif bar.mark.endswith("V1") or "Vert" in bar.notes:
            v_sp = getattr(params, "vert_spacing_in", None)
            v_dim = getattr(params, "wall_length_ft", None)
            if v_sp and v_dim:
                grids.append(HeatmapGrid(
                    mark=bar.mark, size=bar.size,
                    spacing_in=v_sp, dimension_in=v_dim * 12,
                    cover_in=cover, max_spacing_in=max_sp,
                    direction="vertical",
                ))
    return grids


def write_heatmap_to_sheet(ws_validation: Any, grids: list[HeatmapGrid],
                            start_row: int = 4) -> None:
    """
    Write all heatmap grids to the Validation sheet.
    Works with either an xlwings Sheet or openpyxl Worksheet.

    Each grid writes:
      Row 0: mark label + status badge
      Row 1: scale header (position markers every ~12in)
      Rows 2+: visual bar cells (coloured squares)
    """
    row = start_row
    is_xlwings = hasattr(ws_validation, "range")

    for grid in grids:
        positions = grid.bar_positions
        colour    = grid.status_colour
        n_bars    = len(positions)

        # ── Header row ──────────────────────────────────────────────────
        status_txt = {
            "OK":           "✓ Spacing OK",
            "NEAR_MAX":     "⚠ Near ACI max",
            "EXCEEDS_MAX":  "✗ Exceeds ACI max",
        }[grid.status]

        label = (
            f"{grid.mark}  {grid.size}  @{grid.spacing_in:.0f}\" "
            f"{'EF' if 'EF' in (grid.direction or '') else ''}  "
            f"[{n_bars} bars]   {status_txt}"
        )

        if is_xlwings:
            _xw_write_heatmap(ws_validation, grid, row, label, colour)
        else:
            _openpyxl_write_heatmap(ws_validation, grid, row, label, colour)

        row += 6   # each grid is 5 rows tall + 1 spacer

    return row   # return next available row


# ---------------------------------------------------------------------------
# xlwings writer
# ---------------------------------------------------------------------------

def _xw_write_heatmap(ws, grid: HeatmapGrid, start_row: int,
                       label: str, colour: tuple) -> None:
    """Write one heatmap grid using xlwings (live Excel)."""
    positions = grid.bar_positions
    n_bars    = len(positions)
    max_cols  = min(n_bars, 40)   # cap display width at 40 cells

    # Header
    ws.range(f"A{start_row}").value = label
    ws.range(f"A{start_row}").font.bold = True
    ws.range(f"A{start_row}").color = _BLUE

    # Scale row
    scale_row = start_row + 1
    for i in range(max_cols):
        cell = ws.range((scale_row, i + 1))
        cell.value = f'{positions[i]:.0f}"' if i < n_bars else ""
        cell.font.size = 7
        cell.column_width = 4

    # Bar cells — 3 rows tall to look like a wall section
    for vis_row in range(3):
        r = start_row + 2 + vis_row
        for i in range(max_cols):
            cell = ws.range((r, i + 1))
            cell.color = colour if i < n_bars else _WHITE
            cell.row_height = 8

    # Status summary in column after bars
    summary_col = max_cols + 2
    ws.range((start_row + 2, summary_col)).value = (
        f"dim: {grid.dimension_in:.0f}\"  "
        f"cover: {grid.cover_in}\"  "
        f"max_sp: {grid.max_spacing_in:.0f}\""
    )


# ---------------------------------------------------------------------------
# openpyxl writer (used by create_workbook / static export)
# ---------------------------------------------------------------------------

def _openpyxl_write_heatmap(ws, grid: HeatmapGrid, start_row: int,
                              label: str, colour: tuple) -> None:
    """Write one heatmap grid using openpyxl (file-based)."""
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    positions = grid.bar_positions
    n_bars    = len(positions)
    max_cols  = min(n_bars, 40)

    def _fill(rgb):
        return PatternFill("solid", fgColor="{:02X}{:02X}{:02X}".format(*rgb))

    # Header
    hdr = ws.cell(row=start_row, column=1, value=label)
    hdr.font = Font(bold=True, size=10)
    hdr.fill = _fill(_BLUE)

    # Scale row
    for i in range(max_cols):
        c = ws.cell(row=start_row + 1, column=i + 1)
        c.value = f'{positions[i]:.0f}"' if i < n_bars else ""
        c.font  = Font(size=7)
        ws.column_dimensions[get_column_letter(i + 1)].width = 4

    # Bar cells
    for vis_row in range(3):
        r = start_row + 2 + vis_row
        ws.row_dimensions[r].height = 8
        for i in range(max_cols):
            c = ws.cell(row=r, column=i + 1)
            c.fill = _fill(colour if i < n_bars else _WHITE)

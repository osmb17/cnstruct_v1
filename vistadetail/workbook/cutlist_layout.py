"""
CutList tab — static openpyxl placeholder builder.

The live cut-plan content is written by excel_bridge.on_cut_optimize() via xlwings.
This module sets up the initial tab styling so it looks professional before
the optimizer runs.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ── Colours ──────────────────────────────────────────────────────────────────
_NAVY   = "1C3461"
_BLUE   = "2E75B6"
_WHITE  = "FFFFFF"
_GOLD   = "FFD966"
_GRAY   = "F2F2F2"

_thin = Side(style="thin")
_med  = Side(style="medium")


def _fill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)


def _font(size=10, bold=False, color="000000", italic=False) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(l=_thin, r=_thin, t=_thin, b=_thin) -> Border:
    return Border(left=l, right=r, top=t, bottom=b)


# ── Column widths matching the screenshot layout ───────────────────────────
_COL_WIDTHS = {
    "A": 14,   # Stock Bar # / label
    "B": 16,   # Cut 1 / Length
    "C": 16,   # Cut 2 / Use?
    "D": 16,   # Cut 3 / Qty Needed
    "E": 12,   # Used (ft)
    "F": 12,   # Waste (ft)
    "G": 10,   # Waste %
    "H": 8,
    "I": 8,
    "J": 8,
}

_ROW_HEIGHTS = {
    1: 28,   # Title
    2: 20,   # Subtitle
    3: 10,   # spacer
}


def build_cutlist_sheet(ws) -> None:
    """
    Rebuild *ws* (openpyxl Worksheet) as the CutList placeholder.
    Live content is written by on_cut_optimize() at runtime via xlwings.
    """
    # Unmerge any existing merges
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))

    ws.sheet_view.showGridLines = False

    # Column widths
    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # Row heights
    for row, height in _ROW_HEIGHTS.items():
        ws.row_dimensions[row].height = height

    # ── Row 1: Title banner ──────────────────────────────────────────────
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "CUT LIST OPTIMIZER  |  Minimum Waste Bar Cutting Plan (Feature E)"
    c.font      = _font(size=14, bold=True, color=_WHITE)
    c.fill      = _fill(_NAVY)
    c.alignment = _align("center", "center")

    # ── Row 2: Subtitle ──────────────────────────────────────────────────
    ws.merge_cells("A2:J2")
    c = ws["A2"]
    c.value     = "Optimal cut patterns from stock lengths. Minimizes waste and material cost."
    c.font      = _font(size=10, italic=True, color=_WHITE)
    c.fill      = _fill(_BLUE)
    c.alignment = _align("center", "center")

    # ── Row 4: Placeholder message ───────────────────────────────────────
    ws.row_dimensions[4].height = 30
    ws.merge_cells("A4:J4")
    c = ws["A4"]
    c.value     = "Click  CUT OPTIMIZER  on the Dashboard to generate the cut plan."
    c.font      = _font(size=11, italic=True, color="888888")
    c.fill      = _fill(_GRAY)
    c.alignment = _align("center", "center")

    ws.freeze_panes = "A3"

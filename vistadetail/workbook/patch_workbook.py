"""
Patch the existing Rebar Barlist Generator.xlsm in-place.

Preserves all VBA macros and button wiring already set up in the xlsm.
Only rewrites the Dashboard worksheet layout (and optionally BarList,
ReasoningLog, CutList if they need refreshing too).

Run:
    python -m vistadetail.workbook.patch_workbook

No Excel must be open when running this. After running, just open the xlsm.
"""

from __future__ import annotations

import os
import shutil
from datetime import datetime

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

# ---------------------------------------------------------------------------
# Colours (must match create_workbook.py)
# ---------------------------------------------------------------------------
_NAVY        = "1C3461"
_BLUE_MID    = "2E75B6"
_BLUE_LIGHT  = "D6E4F0"
_GREY        = "F2F2F2"
_WHITE       = "FFFFFF"
_ORANGE      = "F4B942"
_GREEN_PALE  = "E8F5E9"
_GREEN_DARK  = "2E7D32"


def _fill(hex_c: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)


def _font(size: int = 10, bold: bool = False, color: str = "000000",
          italic: bool = False) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def _side(style: str = "thin", color: str = "CCCCCC") -> Side:
    return Side(style=style, color=color)


def _border() -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _centre() -> Alignment:
    return _align("center")


# ---------------------------------------------------------------------------
# Safe cell-write helper (handles MergedCell read-only issue)
# ---------------------------------------------------------------------------

def _write(ws, row: int, col: int, value=None,
           font=None, fill=None, alignment=None, border=None) -> None:
    """Write to a plain cell (not a merged stub). Safe against MergedCell."""
    cell = ws.cell(row=row, column=col)
    if hasattr(cell, "value"):
        cell.value = value
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


def _merge_write(ws, cell_range: str, value=None,
                 font=None, fill=None, alignment=None, border=None) -> None:
    """Merge a range then write value + style into the top-left cell."""
    ws.merge_cells(cell_range)
    tl = cell_range.split(":")[0]
    cell = ws[tl]
    cell.value = value
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


# ---------------------------------------------------------------------------
# Dashboard rebuilder
# ---------------------------------------------------------------------------

def _rebuild_dashboard(ws) -> None:
    """
    Completely clear and rebuild the Dashboard worksheet.
    Preserves the sheet object (and therefore any button shapes already on it).
    """
    from vistadetail.engine.templates import TEMPLATE_NAMES
    from openpyxl.worksheet.datavalidation import DataValidation

    # ── 1. Unmerge everything first (avoids MergedCell write errors) ─────
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))

    # ── 2. Clear all cell content + styles ───────────────────────────────
    for row in ws.iter_rows():
        for cell in row:
            cell.value     = None
            cell.font      = Font()
            cell.fill      = PatternFill()
            cell.border    = Border()
            cell.alignment = Alignment()

    # ── 3. Column widths ─────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16
    ws.sheet_view.showGridLines = False

    # ── 4. Row 1: Title ──────────────────────────────────────────────────
    ws.row_dimensions[1].height = 40
    _merge_write(ws, "A1:E1",
                 value="VISTA REBAR BARLIST GENERATOR",
                 font=_font(16, bold=True, color=_WHITE),
                 fill=_fill(_NAVY),
                 alignment=_centre())

    # ── Row 2: (removed blue subtitle — now blank spacer) ─────────────────
    ws.row_dimensions[2].height = 6

    # ── Rows 3-5: Input fields ────────────────────────────────────────────
    fields = [
        (3, "Structure Type:",  TEMPLATE_NAMES[0]),
        (4, "Project No:",      ""),
        (5, "Drawn By:",        ""),
    ]
    for row, label, default in fields:
        ws.row_dimensions[row].height = 22
        _write(ws, row, 1, label,
               font=_font(10, bold=True),
               fill=_fill(_GREY), border=_border(),
               alignment=_align("right"))
        _write(ws, row, 2, default,
               font=_font(10),
               fill=_fill(_BLUE_LIGHT), border=_border(),
               alignment=_align("left"))

    # Template dropdown on B3.
    # Mac Excel does NOT reliably support cross-sheet references in data
    # validation. Store the list in column G (hidden) on this same sheet —
    # same-sheet references always work. Column G rows 3..N+2.
    n = len(TEMPLATE_NAMES)
    for i, name in enumerate(TEMPLATE_NAMES):
        _write(ws, i + 3, 7, name)   # G3, G4, ... G(N+2)
    ws.column_dimensions["G"].hidden = True

    dv = DataValidation(
        type="list",
        formula1=f"$G$3:$G${n + 2}",
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Template",
        error="Please pick a template from the dropdown list.",
    )
    dv.sqref = "B3"
    ws.add_data_validation(dv)
    _write(ws, 3, 3, "← pick template, then Refresh Inputs",
           font=_font(9, italic=True, color="888888"))

    # ── Row 6: Spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 8

    # ── Row 7: Primary action buttons ────────────────────────────────────
    ws.row_dimensions[7].height = 30
    btn_labels = ["GENERATE DRAFT", "REFRESH INPUTS", "CLEAR ALL", "EXPORT CSV"]
    btn_fills  = [_ORANGE, _BLUE_MID, _BLUE_MID, _BLUE_MID]
    for col, (label, bg) in enumerate(zip(btn_labels, btn_fills), start=1):
        _write(ws, 7, col, label,
               font=_font(10, bold=True, color=_WHITE),
               fill=_fill(bg), border=_border(),
               alignment=_centre())

    # ── Row 8: Status ─────────────────────────────────────────────────────
    ws.row_dimensions[8].height = 22
    _write(ws, 8, 1, "Status:",
           font=_font(10, bold=True),
           fill=_fill(_GREY), border=_border())
    _write(ws, 8, 2, "✓ Ready",
           font=_font(10, color=_GREEN_DARK),
           fill=_fill("F1F8E9"), border=_border())

    # ── Row 9: Spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[9].height = 8

    # ── Row 10: Results header ─────────────────────────────────────────────
    ws.row_dimensions[10].height = 22
    _merge_write(ws, "A10:E10",
                 value="LAST GENERATION RESULTS",
                 font=_font(9, bold=True, color=_WHITE),
                 fill=_fill(_NAVY),
                 alignment=_centre())

    # ── Row 11: Total bars + weight ────────────────────────────────────────
    ws.row_dimensions[11].height = 24
    _write(ws, 11, 1, "Total Bars:",
           font=_font(10, bold=True), fill=_fill(_GREY), border=_border())
    _write(ws, 11, 2, "—",
           font=_font(13, bold=True, color=_GREEN_DARK),
           fill=_fill(_GREEN_PALE), border=_border(),
           alignment=_centre())
    _write(ws, 11, 3, "Total Weight (lb):",
           font=_font(10, bold=True), fill=_fill(_GREY), border=_border())
    _write(ws, 11, 4, "—",
           font=_font(13, bold=True, color=_GREEN_DARK),
           fill=_fill(_GREEN_PALE), border=_border(),
           alignment=_centre())

    # ── Row 12: Cost + rate ────────────────────────────────────────────────
    ws.row_dimensions[12].height = 24
    _write(ws, 12, 1, "Est. Material Cost:",
           font=_font(10, bold=True), fill=_fill(_GREY), border=_border())
    _write(ws, 12, 2, "—",
           font=_font(13, bold=True, color=_GREEN_DARK),
           fill=_fill(_GREEN_PALE), border=_border(),
           alignment=_centre())
    _write(ws, 12, 3, "Rate ($/lb):",
           font=_font(10, bold=True), fill=_fill(_GREY), border=_border())
    _write(ws, 12, 4, 0.80,
           font=_font(10),
           fill=_fill("FFF9C4"),   # pale yellow = user editable
           border=_border(),
           alignment=_centre())
    _write(ws, 12, 5, "← change rate then re-generate",
           font=_font(8, italic=True, color="888888"))

    # ── Row 13: Spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[13].height = 8

    # ── Row 14: Confidence / acceptance rate ────────────────────────────────
    ws.row_dimensions[14].height = 20
    # (populated at runtime by on_generate via _write_confidence)

    # ── Row 15: Spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[15].height = 8

    # ── Row 16: Secondary buttons ─────────────────────────────────────────
    ws.row_dimensions[16].height = 26
    sec = ["LOG CORRECTIONS", "CUT OPTIMIZER", "COMPOSE PROJECT", "SAVE GOLD", "CLEAR GOLD"]
    for col, label in enumerate(sec, start=1):
        _write(ws, 16, col, label,
               font=_font(9, bold=True, color=_WHITE),
               fill=_fill(_BLUE_MID), border=_border(),
               alignment=_centre())


# ---------------------------------------------------------------------------
# Main patch function
# ---------------------------------------------------------------------------

def patch(xlsm_path: str | None = None) -> str:
    """
    Open the existing xlsm with keep_vba=True, rebuild the Dashboard,
    and save in place. All VBA / macros are preserved.
    """
    if xlsm_path is None:
        base = os.path.abspath(os.path.join(os.path.dirname(__file__), "..", ".."))
        for name in ("Rebar Barlist Generator.xlsm",):
            p = os.path.join(base, name)
            if os.path.exists(p):
                xlsm_path = p
                break
        if xlsm_path is None:
            raise FileNotFoundError(
                "Rebar Barlist Generator.xlsm not found in RebarGenerator folder."
            )

    # Back up before patching
    ts  = datetime.now().strftime("%Y%m%d_%H%M%S")
    bak = xlsm_path + f".bak_{ts}"
    shutil.copy2(xlsm_path, bak)
    print(f"  Backup: {bak}")

    print(f"  Opening (keep_vba=True): {xlsm_path}")
    wb = openpyxl.load_workbook(xlsm_path, keep_vba=True)

    # Rebuild Dashboard
    if "Dashboard" not in wb.sheetnames:
        wb.create_sheet("Dashboard", 0)
    ws_dash = wb["Dashboard"]
    _rebuild_dashboard(ws_dash)
    print("  ✓ Dashboard rebuilt")

    # Also rebuild BarList (fixes any merge issues there too)
    if "BarList" in wb.sheetnames:
        from vistadetail.workbook.barlist_layout import build_barlist_sheet
        build_barlist_sheet(wb["BarList"])
        print("  ✓ BarList rebuilt")

    # Also rebuild ReasoningLog
    if "ReasoningLog" in wb.sheetnames:
        from vistadetail.workbook.reasoning_layout import build_reasoning_sheet
        build_reasoning_sheet(wb["ReasoningLog"])
        print("  ✓ ReasoningLog rebuilt")

    # Also rebuild CutList
    if "CutList" in wb.sheetnames:
        from vistadetail.workbook.cutlist_layout import build_cutlist_sheet
        build_cutlist_sheet(wb["CutList"])
        print("  ✓ CutList rebuilt")

    # Rebuild _Templates hidden sheet.
    # Layout (rows are re-written by SetupDropdowns VBA on every Workbook_Open):
    #   Rows 1..N   — template names for the Dashboard B3 dropdown (col A)
    #   Row  N+1    — blank separator
    #   Row  N+2    — "COMPOSE PROJECT INPUT AREA" heading
    #   Row  N+3    — column header: A=Prefix | B=Template | C=Label | D/E=key/val …
    #   Row  N+4    — hint row
    #   Row  N+5+   — user data (one row per structure to compose)
    from vistadetail.engine.templates import TEMPLATE_NAMES
    n = len(TEMPLATE_NAMES)
    if "_Templates" not in wb.sheetnames:
        ws_tmpl = wb.create_sheet("_Templates")
    else:
        ws_tmpl = wb["_Templates"]
    ws_tmpl.sheet_state = "hidden"
    # Clear all existing values
    for row in ws_tmpl.iter_rows():
        for cell in row:
            cell.value = None
    # Rows 1..N — template names (also written by VBA SetupDropdowns on open)
    for i, name in enumerate(TEMPLATE_NAMES, start=1):
        ws_tmpl.cell(row=i, column=1, value=name)
    # Blank separator
    # Row N+2 — compose section header
    sep = n + 2
    ws_tmpl.cell(row=sep, column=1, value="COMPOSE PROJECT INPUT AREA")
    ws_tmpl.cell(row=sep, column=1).font = _font(10, bold=True)
    # Row N+3 — column headers
    hdr = sep + 1
    for col, label in enumerate(
        ["Prefix", "Template", "Label", "param_key_1", "param_val_1", "param_key_2", "param_val_2"],
        start=1,
    ):
        ws_tmpl.cell(row=hdr, column=col, value=label)
        ws_tmpl.cell(row=hdr, column=col).font = _font(9, bold=True)
        ws_tmpl.cell(row=hdr, column=col).fill = _fill(_GREY)
    # Row N+4 — example / hint row
    hint = hdr + 1
    ex = ["HW", "Straight Headwall", "Headwall – Sta 12+50", "wall_width_ft", 8.0, "wall_height_ft", 5.0]
    for col, val in enumerate(ex, start=1):
        ws_tmpl.cell(row=hint, column=col, value=val)
        ws_tmpl.cell(row=hint, column=col).font = _font(9, italic=True, color="888888")
    print(f"  ✓ _Templates rebuilt ({n} templates + compose header at row {sep})")

    wb.save(xlsm_path)
    print(f"\n✓ Patched and saved: {xlsm_path}")
    print("  Open the file in Excel — no 'Repaired' message, all buttons intact.")
    return xlsm_path


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else None
    patch(path)

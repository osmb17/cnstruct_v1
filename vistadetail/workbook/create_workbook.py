"""
Create the Rebar Barlist Generator.xlsm workbook scaffold via openpyxl.

Run once (or after any layout change) to regenerate the workbook:
    cd /Users/osmb/VistaSteel/RebarGenerator
    python -m vistadetail.workbook.create_workbook

Then open the .xlsx in Excel → Save As → Excel Macro-Enabled (.xlsm)
Then run:  python -m vistadetail.setup_xlwings
"""

from __future__ import annotations

import os
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ---------------------------------------------------------------------------
# Palette
# ---------------------------------------------------------------------------
_NAVY        = "1C3461"   # primary dark navy
_BLUE_MID    = "2E75B6"   # medium blue (section headers)
_BLUE_LIGHT  = "D6E4F0"   # light blue (input fields)
_GREY        = "F2F2F2"   # label background
_WHITE       = "FFFFFF"
_ORANGE      = "F4B942"   # GENERATE button accent
_GREEN_PALE  = "E8F5E9"   # results background
_GREEN_DARK  = "2E7D32"   # results text colour

WORKBOOK_PATH = os.path.abspath(
    os.path.join(os.path.dirname(__file__), "..", "..", "Rebar Barlist Generator.xlsx")
)


# ---------------------------------------------------------------------------
# Style helpers
# ---------------------------------------------------------------------------

def _fill(hex_c: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)


def _font(size: int = 10, bold: bool = False, color: str = "000000",
          italic: bool = False, name: str = "Calibri") -> Font:
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def _side(style: str = "thin", color: str = "CCCCCC") -> Side:
    return Side(style=style, color=color)


def _border(all_thin: bool = True) -> Border:
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s) if all_thin else Border()


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _centre() -> Alignment:
    return _align("center")


def _merge_write(ws, cell_range: str, value="", font=None, fill=None,
                 alignment=None, border=None) -> None:
    """Merge a range, write value + style into top-left cell only."""
    ws.merge_cells(cell_range)
    tl = cell_range.split(":")[0]
    cell = ws[tl]
    cell.value = value
    if font:      cell.font      = font
    if fill:      cell.fill      = fill
    if alignment: cell.alignment = alignment
    if border:    cell.border    = border


# ---------------------------------------------------------------------------
# Dashboard tab
# ---------------------------------------------------------------------------

def _build_dashboard(ws) -> None:
    ws.title = "Dashboard"
    ws.sheet_view.showGridLines = False

    # Column widths
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 28
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    # ── Row 1: Title banner ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 40
    _merge_write(ws, "A1:E1",
                 value="VISTA REBAR BARLIST GENERATOR",
                 font=_font(16, bold=True, color=_WHITE),
                 fill=_fill(_NAVY),
                 alignment=_centre())

    # ── Row 2: Subtitle (company / standards) ────────────────────────────
    ws.row_dimensions[2].height = 18
    _merge_write(ws, "A2:E2",
                 value="ACI 318-19  |  Caltrans BDS  |  Vista Steel Company, Ventura CA",
                 font=_font(9, italic=True, color=_WHITE),
                 fill=_fill(_BLUE_MID),
                 alignment=_centre())

    # ── Rows 3-5: Input fields ────────────────────────────────────────────
    from vistadetail.engine.templates import TEMPLATE_NAMES
    fields = [
        (3, "Structure Type:",  TEMPLATE_NAMES[0]),
        (4, "Project No:",      ""),
        (5, "Drawn By:",        ""),
    ]
    for row, label, default in fields:
        ws.row_dimensions[row].height = 22
        lc = ws.cell(row=row, column=1, value=label)
        lc.font      = _font(10, bold=True)
        lc.fill      = _fill(_GREY)
        lc.border    = _border()
        lc.alignment = _align("right")
        vc = ws.cell(row=row, column=2, value=default)
        vc.font      = _font(10)
        vc.fill      = _fill(_BLUE_LIGHT)
        vc.border    = _border()
        vc.alignment = _align("left")

    # Template dropdown on B3
    template_list = ",".join(TEMPLATE_NAMES)
    dv = DataValidation(
        type="list",
        formula1=f'"{template_list}"',
        allow_blank=False,
        showDropDown=False,
        showErrorMessage=True,
        errorTitle="Invalid Template",
        error="Please pick a template from the dropdown list.",
    )
    dv.sqref = "B3"
    ws.add_data_validation(dv)
    ws.cell(row=3, column=3).value = "← pick template, then Refresh Inputs"
    ws.cell(row=3, column=3).font  = _font(9, italic=True, color="888888")

    # ── Row 6: Spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[6].height = 8

    # ── Row 7: Primary buttons ────────────────────────────────────────────
    ws.row_dimensions[7].height = 30
    btns = [
        (1, "GENERATE DRAFT",  _ORANGE),
        (2, "REFRESH INPUTS",  _BLUE_MID),
        (3, "CLEAR ALL",       _BLUE_MID),
        (4, "EXPORT CSV",      _BLUE_MID),
    ]
    for col, label, bg in btns:
        c = ws.cell(row=7, column=col, value=label)
        c.font      = _font(10, bold=True, color=_WHITE)
        c.fill      = _fill(bg)
        c.alignment = _centre()
        c.border    = _border()

    # ── Row 8: Status line ────────────────────────────────────────────────
    ws.row_dimensions[8].height = 22
    sc = ws.cell(row=8, column=1, value="Status:")
    sc.font      = _font(10, bold=True)
    sc.fill      = _fill(_GREY)
    sc.border    = _border()
    sv = ws.cell(row=8, column=2, value="✓ Ready")
    sv.font      = _font(10, color=_GREEN_DARK)
    sv.fill      = _fill("F1F8E9")
    sv.border    = _border()

    # ── Row 9: Spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[9].height = 8

    # ── Row 10: Results header ─────────────────────────────────────────────
    ws.row_dimensions[10].height = 22
    _merge_write(ws, "A10:E10",
                 value="LAST GENERATION RESULTS",
                 font=_font(9, bold=True, color=_WHITE),
                 fill=_fill(_NAVY),
                 alignment=_centre())

    # ── Row 11: Bar count + weight ─────────────────────────────────────────
    ws.row_dimensions[11].height = 22
    ws.cell(row=11, column=1, value="Total Bars:").font   = _font(10, bold=True)
    ws.cell(row=11, column=1).fill    = _fill(_GREY)
    ws.cell(row=11, column=1).border  = _border()
    ws.cell(row=11, column=2, value="-").font  = _font(11, bold=True, color=_GREEN_DARK)
    ws.cell(row=11, column=2).fill   = _fill(_GREEN_PALE)
    ws.cell(row=11, column=2).border = _border()
    ws.cell(row=11, column=2).alignment = _align("center")
    ws.cell(row=11, column=3, value="Total Weight (lb):").font  = _font(10, bold=True)
    ws.cell(row=11, column=3).fill   = _fill(_GREY)
    ws.cell(row=11, column=3).border = _border()
    ws.cell(row=11, column=4, value="-").font  = _font(11, bold=True, color=_GREEN_DARK)
    ws.cell(row=11, column=4).fill   = _fill(_GREEN_PALE)
    ws.cell(row=11, column=4).border = _border()
    ws.cell(row=11, column=4).alignment = _align("center")

    # ── Row 12: Material cost ─────────────────────────────────────────────
    ws.row_dimensions[12].height = 22
    ws.cell(row=12, column=1, value="Est. Material Cost:").font = _font(10, bold=True)
    ws.cell(row=12, column=1).fill    = _fill(_GREY)
    ws.cell(row=12, column=1).border  = _border()
    ws.cell(row=12, column=2, value="-").font  = _font(11, bold=True, color=_GREEN_DARK)
    ws.cell(row=12, column=2).fill   = _fill(_GREEN_PALE)
    ws.cell(row=12, column=2).border = _border()
    ws.cell(row=12, column=2).alignment = _align("center")
    ws.cell(row=12, column=3, value="Rate ($/lb):").font  = _font(10, bold=True)
    ws.cell(row=12, column=3).fill   = _fill(_GREY)
    ws.cell(row=12, column=3).border = _border()
    ws.cell(row=12, column=4, value=0.80).font  = _font(10)   # user-editable
    ws.cell(row=12, column=4).fill   = _fill("FFF9C4")        # pale yellow = editable
    ws.cell(row=12, column=4).border = _border()
    ws.cell(row=12, column=4).alignment = _align("center")
    ws.cell(row=12, column=5, value="← change rate then re-generate").font = \
        _font(8, italic=True, color="888888")

    # ── Row 13: Spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[13].height = 8

    # ── Row 14: Acceptance / confidence ────────────────────────────────────
    ws.row_dimensions[14].height = 20
    ws.cell(row=14, column=1, value="").font = _font(9, italic=True, color=_GREEN_DARK)

    # ── Row 15: Spacer ─────────────────────────────────────────────────────
    ws.row_dimensions[15].height = 8

    # ── Row 16: Secondary buttons ─────────────────────────────────────────
    ws.row_dimensions[16].height = 26
    sec_btns = [
        (1, "LOG CORRECTIONS", _BLUE_MID),
        (2, "CUT OPTIMIZER",   _BLUE_MID),
        (3, "COMPOSE PROJECT", _BLUE_MID),
        (4, "SAVE GOLD",       _BLUE_MID),
        (5, "CLEAR GOLD",      _BLUE_MID),
    ]
    for col, label, bg in sec_btns:
        c = ws.cell(row=16, column=col, value=label)
        c.font      = _font(9, bold=True, color=_WHITE)
        c.fill      = _fill(bg)
        c.alignment = _centre()
        c.border    = _border()


# ---------------------------------------------------------------------------
# Inputs tab
# ---------------------------------------------------------------------------

def _build_inputs(ws) -> None:
    ws.title = "Inputs"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 32

    ws.row_dimensions[1].height = 28
    _merge_write(ws, "A1:C1", value="Template Inputs",
                 font=_font(12, bold=True, color=_WHITE),
                 fill=_fill(_NAVY), alignment=_centre())

    ws.row_dimensions[2].height = 16
    _merge_write(ws, "A2:C2",
                 value="← populated automatically when you click Refresh Inputs",
                 font=_font(9, italic=True, color="888888"),
                 alignment=_align("center"))

    # Pre-populate with first template
    from vistadetail.engine.templates import TEMPLATE_REGISTRY, TEMPLATE_NAMES
    tmpl = TEMPLATE_REGISTRY[TEMPLATE_NAMES[0]]
    for i, field in enumerate(tmpl.inputs):
        row = i + 3
        ws.row_dimensions[row].height = 20
        lc = ws.cell(row=row, column=1, value=field.label or field.name)
        lc.font      = _font(10)
        lc.fill      = _fill(_GREY if i % 2 == 0 else _WHITE)
        lc.border    = _border()
        vc = ws.cell(row=row, column=2, value=field.default)
        vc.font      = _font(10)
        vc.fill      = _fill(_BLUE_LIGHT)
        vc.border    = _border()
        vc.alignment = _align("center")
        if field.hint:
            hc = ws.cell(row=row, column=3, value=f"  {field.hint}")
            hc.font = _font(9, italic=True, color="888888")


# ---------------------------------------------------------------------------
# BarList tab  (Vista Steel shop drawing format)
# ---------------------------------------------------------------------------

def _build_barlist(ws) -> None:
    ws.title = "BarList"
    from vistadetail.workbook.barlist_layout import build_barlist_sheet
    build_barlist_sheet(ws)


# ---------------------------------------------------------------------------
# ReasoningLog tab
# ---------------------------------------------------------------------------

def _build_reasoning_log(ws) -> None:
    ws.title = "ReasoningLog"
    from vistadetail.workbook.reasoning_layout import build_reasoning_sheet
    build_reasoning_sheet(ws)


# ---------------------------------------------------------------------------
# CutList tab
# ---------------------------------------------------------------------------

def _build_cutlist(ws) -> None:
    ws.title = "CutList"
    from vistadetail.workbook.cutlist_layout import build_cutlist_sheet
    build_cutlist_sheet(ws)


# ---------------------------------------------------------------------------
# Validation tab
# ---------------------------------------------------------------------------

def _build_validation(ws) -> None:
    ws.title = "Validation"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 70

    ws.row_dimensions[1].height = 28
    _merge_write(ws, "A1:B1", value="Warnings & Flags",
                 font=_font(12, bold=True, color=_WHITE),
                 fill=_fill(_NAVY), alignment=_centre())

    for i, h in enumerate(["Severity", "Message"], start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font      = _font(10, bold=True)
        c.fill      = _fill(_GREY)
        c.border    = _border()
        c.alignment = _align("center")

    ws.freeze_panes = "A3"


# ---------------------------------------------------------------------------
# Hidden _Templates tab
# ---------------------------------------------------------------------------

def _build_templates_hidden(ws) -> None:
    ws.title = "_Templates"
    ws.sheet_state = "hidden"
    ws["A1"].value = "Template definitions are managed in Python. Do not edit."
    ws["A1"].font  = _font(9, italic=True, color="888888")


# ---------------------------------------------------------------------------
# Main entry point
# ---------------------------------------------------------------------------

def create_workbook(path: str = WORKBOOK_PATH) -> str:
    """
    Build the complete Rebar Barlist Generator.xlsx from scratch.
    Returns the saved path.

    Workflow after running this:
      1. Open Rebar Barlist Generator.xlsx in Excel
      2. File → Save As → Excel Macro-Enabled Workbook (.xlsm)
         (keep same filename, just change extension)
      3. python -m vistadetail.setup_xlwings
    """
    wb = Workbook()
    wb.remove(wb.active)   # remove default blank sheet

    _build_dashboard(wb.create_sheet("Dashboard"))
    _build_inputs(wb.create_sheet("Inputs"))
    _build_barlist(wb.create_sheet("BarList"))
    _build_reasoning_log(wb.create_sheet("ReasoningLog"))
    _build_validation(wb.create_sheet("Validation"))
    _build_cutlist(wb.create_sheet("CutList"))
    _build_templates_hidden(wb.create_sheet("_Templates"))

    # openpyxl can only produce valid .xlsx; user must re-save as .xlsm in Excel
    if path.endswith(".xlsm"):
        path = path[:-5] + ".xlsx"

    wb.save(path)
    print(f"\n✓ Workbook saved: {path}")
    print("\nNext steps:")
    print("  1. Open the file in Excel")
    print("  2. File → Save As → Excel Macro-Enabled Workbook (.xlsm)")
    print("     (keep the same filename, just change the extension)")
    print("  3. python -m vistadetail.setup_xlwings")
    return path


if __name__ == "__main__":
    create_workbook()

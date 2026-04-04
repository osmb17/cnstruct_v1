"""
Vista Steel Shop Drawing Barlist Layout — openpyxl builder.

Builds the BarList worksheet to match the Vista Steel paper barlist format:

  Rows 1-8:   Header (company info, job fields, delivery date, etc.)
  Row  9:     Column headers (S/H/L | NO.OF UNITS | NO.PER UNIT | TOTAL |
                              SIZE | GRADE | LENGTH | MARK | TYPE | A..H | #)
  Rows 10-69: Data rows (60 bar entries, row-numbered 1-60 on right)

Column layout (A–R):
  A  S/H/L         B  NO. OF UNITS   C  NO. PER UNIT  D  TOTAL
  E  SIZE           F  GRADE          G  LENGTH         H  MARK
  I  TYPE           J  A              K  B              L  C'
  M  D              N  E              O  F              P  G
  Q  H              R  [row number]

Used by create_workbook.py and by the workbook rebuild script.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

HEADER_ROWS   = 5            # rows 1-5: job info header
COL_HDR_ROW   = 6            # row 6: column labels
DATA_START    = 7            # row 7: first bar data row
DATA_ROWS     = 60           # support up to 60 marks per sheet
DATA_END      = DATA_START + DATA_ROWS - 1   # row 66

# Column indices (1-based)
COL_SHL       = 1   # A
COL_UNITS     = 2   # B
COL_PER_UNIT  = 3   # C
COL_TOTAL     = 4   # D
COL_SIZE      = 5   # E
COL_GRADE     = 6   # F
COL_LENGTH    = 7   # G
COL_MARK      = 8   # H
COL_TYPE      = 9   # I
COL_A         = 10  # J
COL_B         = 11  # K
COL_C         = 12  # L
COL_D         = 13  # M
COL_E         = 14  # N
COL_F         = 15  # O
COL_G         = 16  # P
COL_H         = 17  # Q
COL_ROWNUM    = 18  # R

# Shape → S/H/L indicator
SHL_MAP = {
    "Str":    "S",
    "L":      "L",
    "U":      "U",
    "Hook":   "H",
    "Hoop":   "H",
    "Spiral": "S",
}

# Shape → bend type number / code
TYPE_MAP = {
    "Str":    "0",
    "L":      "1",
    "U":      "2",
    "Hook":   "1",
    "Hoop":   "SQ",
    "Spiral": "SPI",
}

# ---------------------------------------------------------------------------
# Colour / style helpers
# ---------------------------------------------------------------------------

_NAVY   = "1C3461"   # Vista Steel dark navy (header)
_WHITE  = "FFFFFF"
_GRAY   = "F0F0F0"   # label background
_LBLUE  = "EBF0FA"   # fillable field background
_ROW_A  = "FAFAFA"   # alternating row tint
_ROW_B  = "FFFFFF"


def _fill(hex_c: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_c)


def _font(size: int = 9, bold: bool = False, color: str = "000000",
          name: str = "Calibri", italic: bool = False) -> Font:
    return Font(name=name, size=size, bold=bold, color=color, italic=italic)


def _align(h: str = "left", v: str = "center", wrap: bool = False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


_thin  = Side(style="thin")
_med   = Side(style="medium")
_thick = Side(style="thick")
_white_med = Side(style="medium", color=_WHITE)


def _border(l=_thin, r=_thin, t=_thin, b=_thin) -> Border:
    return Border(left=l, right=r, top=t, bottom=b)


# ---------------------------------------------------------------------------
# Main builder
# ---------------------------------------------------------------------------

def build_barlist_sheet(ws) -> None:
    """
    Rebuild *ws* (an openpyxl Worksheet) as a Vista Steel barlist.
    Existing content is cleared first.
    """
    # ── Clear ────────────────────────────────────────────────────────────
    # Must unmerge first — writing to MergedCell raises AttributeError
    for merge in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merge))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None
            cell.font      = Font()
            cell.fill      = PatternFill()
            cell.border    = Border()
            cell.alignment = Alignment()

    # ── Column widths ────────────────────────────────────────────────────
    col_widths = {
        "A": 6.0,  "B": 9.5,  "C": 9.5,  "D": 7.5,
        "E": 7.0,  "F": 7.0,  "G": 12.0, "H": 9.0,
        "I": 7.0,  "J": 7.0,  "K": 7.0,  "L": 7.0,
        "M": 7.0,  "N": 7.0,  "O": 7.0,  "P": 7.0,
        "Q": 7.0,  "R": 4.5,
    }
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    # ── Row heights ──────────────────────────────────────────────────────
    for r in range(1, HEADER_ROWS + 1):
        ws.row_dimensions[r].height = 22   # header info rows
    ws.row_dimensions[COL_HDR_ROW].height = 30  # column label row
    for r in range(DATA_START, DATA_END + 1):
        ws.row_dimensions[r].height = 18   # data rows

    # ── HEADER SECTION ───────────────────────────────────────────────────
    _build_header(ws)

    # ── COLUMN HEADERS (row 9) ───────────────────────────────────────────
    _build_col_headers(ws)

    # ── DATA ROWS (10-69) ────────────────────────────────────────────────
    _build_data_rows(ws)

    # ── Freeze panes below header ────────────────────────────────────────
    ws.freeze_panes = ws.cell(row=DATA_START, column=1)


# ---------------------------------------------------------------------------
# Header section
# ---------------------------------------------------------------------------

def _set(ws, row: int, col: int | str, value="",
         fnt=None, fil=None, brd=None, aln=None) -> None:
    """Set a single cell's properties."""
    if isinstance(col, str):
        col = ord(col) - ord("A") + 1
    cell = ws.cell(row=row, column=col)
    cell.value = value
    if fnt: cell.font      = fnt
    if fil: cell.fill      = fil
    if brd: cell.border    = brd
    if aln: cell.alignment = aln
    return cell


def _merge_set(ws, top_left: str, bottom_right: str, value="",
               fnt=None, fil=None, brd=None, aln=None) -> None:
    """Merge a range then set the top-left cell."""
    ws.merge_cells(f"{top_left}:{bottom_right}")
    cell = ws[top_left]
    cell.value = value
    if fnt: cell.font      = fnt
    if fil: cell.fill      = fil
    if brd: cell.border    = brd
    if aln: cell.alignment = aln


def _build_header(ws) -> None:
    # ── Row 1: DELIVERY DATE | JOB # | LOT # ─────────────────────────────
    ws.merge_cells("A1:C1")
    ws["A1"].value     = "DELIVERY DATE:"
    ws["A1"].font      = _font(size=10, bold=True)
    ws["A1"].fill      = _fill(_GRAY)
    ws["A1"].alignment = _align("right", "center")

    ws.merge_cells("D1:H1")
    ws["D1"].fill   = _fill(_LBLUE)
    ws["D1"].border = _border(b=_med)

    ws["I1"].value     = "JOB #"
    ws["I1"].font      = _font(size=10, bold=True)
    ws["I1"].fill      = _fill(_GRAY)
    ws["I1"].alignment = _align("right", "center")

    ws.merge_cells("J1:L1")
    ws["J1"].fill   = _fill(_LBLUE)
    ws["J1"].border = _border(b=_med)

    ws["M1"].value     = "LOT #"
    ws["M1"].font      = _font(size=10, bold=True)
    ws["M1"].fill      = _fill(_GRAY)
    ws["M1"].alignment = _align("right", "center")

    ws.merge_cells("N1:R1")
    ws["N1"].fill   = _fill(_LBLUE)
    ws["N1"].border = _border(b=_med)

    # ── Row 2: JOB (description) | COLOR ─────────────────────────────────
    ws.merge_cells("A2:B2")
    ws["A2"].value     = "JOB"
    ws["A2"].font      = _font(size=10, bold=True)
    ws["A2"].fill      = _fill(_GRAY)
    ws["A2"].alignment = _align("right", "center")
    ws["A2"].border    = _border(b=_med)

    ws.merge_cells("C2:H2")
    ws["C2"].fill   = _fill(_LBLUE)
    ws["C2"].border = _border(b=_med)

    ws["I2"].value     = "COLOR"
    ws["I2"].font      = _font(size=10, bold=True)
    ws["I2"].fill      = _fill(_GRAY)
    ws["I2"].alignment = _align("right", "center")

    ws.merge_cells("J2:R2")
    ws["J2"].fill   = _fill(_LBLUE)
    ws["J2"].border = _border(b=_med)

    # ── Row 3: CONTRACTOR | DATE ──────────────────────────────────────────
    ws.merge_cells("A3:B3")
    ws["A3"].value     = "CONTRACTOR"
    ws["A3"].font      = _font(size=10, bold=True)
    ws["A3"].fill      = _fill(_GRAY)
    ws["A3"].alignment = _align("right", "center")
    ws["A3"].border    = _border(b=_med)

    ws.merge_cells("C3:H3")
    ws["C3"].fill   = _fill(_LBLUE)
    ws["C3"].border = _border(b=_med)

    ws["I3"].value     = "DATE"
    ws["I3"].font      = _font(size=10, bold=True)
    ws["I3"].fill      = _fill(_GRAY)
    ws["I3"].alignment = _align("right", "center")

    ws.merge_cells("J3:R3")
    ws["J3"].fill   = _fill(_LBLUE)
    ws["J3"].border = _border(b=_med)

    # ── Row 4: MATERIAL FOR | CODE | DETAILER ────────────────────────────
    ws.merge_cells("A4:B4")
    ws["A4"].value     = "MATERIAL FOR"
    ws["A4"].font      = _font(size=10, bold=True)
    ws["A4"].fill      = _fill(_GRAY)
    ws["A4"].alignment = _align("right", "center")
    ws["A4"].border    = _border(b=_med)

    ws.merge_cells("C4:F4")
    ws["C4"].fill   = _fill(_LBLUE)
    ws["C4"].border = _border(b=_med)

    ws["G4"].value     = "CODE"
    ws["G4"].font      = _font(size=10, bold=True)
    ws["G4"].fill      = _fill(_GRAY)
    ws["G4"].alignment = _align("right", "center")

    ws.merge_cells("H4:I4")
    ws["H4"].fill   = _fill(_LBLUE)
    ws["H4"].border = _border(b=_med)

    ws["J4"].value     = "DETAILER"
    ws["J4"].font      = _font(size=10, bold=True)
    ws["J4"].fill      = _fill(_GRAY)
    ws["J4"].alignment = _align("right", "center")

    ws.merge_cells("K4:R4")
    ws["K4"].fill   = _fill(_LBLUE)
    ws["K4"].border = _border(b=_med)

    # ── Row 5: FOB / INSTALLED / ASSEMBLED checkboxes ─────────────────────
    ws.merge_cells("A5:F5")
    ws["A5"].value     = "\u25a1 FOB"
    ws["A5"].font      = _font(size=10, bold=True)
    ws["A5"].alignment = _align("center", "center")

    ws.merge_cells("G5:L5")
    ws["G5"].value     = "\u25a1 INSTALLED"
    ws["G5"].font      = _font(size=10, bold=True)
    ws["G5"].alignment = _align("center", "center")

    ws.merge_cells("M5:R5")
    ws["M5"].value     = "\u25a1 ASSEMBLED"
    ws["M5"].font      = _font(size=10, bold=True)
    ws["M5"].alignment = _align("center", "center")

    # Heavy separator line across row 5 bottom (last header row)
    for col in range(1, 19):
        cell = ws.cell(row=5, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=cell.border.top,
            bottom=_thick,
        )


# ---------------------------------------------------------------------------
# Column header row (row 9)
# ---------------------------------------------------------------------------

_COL_HEADERS = [
    "S/H/L",
    "NO. OF\nUNITS",
    "NO. PER\nUNIT",
    "TOTAL",
    "SIZE",
    "GRADE",
    "LENGTH",
    "MARK",
    "TYPE",
    "A",
    "B",
    "C'",
    "D",
    "E",
    "F",
    "G",
    "H",
    "",        # row number column (blank header)
]


def _build_col_headers(ws) -> None:
    row = COL_HDR_ROW
    for i, label in enumerate(_COL_HEADERS):
        col = i + 1
        cell = ws.cell(row=row, column=col, value=label)
        cell.font      = Font(name="Calibri", size=9, bold=True, color=_WHITE)
        cell.fill      = _fill(_NAVY)
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=True)
        # Left edge thick, right edge thick, inner medium-white dividers
        l = _thick if col == 1            else _white_med
        r = _thick if col == COL_ROWNUM  else _white_med
        cell.border = Border(left=l, right=r, top=_thick, bottom=_med)


# ---------------------------------------------------------------------------
# Data rows (10-69)
# ---------------------------------------------------------------------------

def _build_data_rows(ws) -> None:
    for i in range(DATA_ROWS):
        row     = DATA_START + i
        row_num = i + 1
        even    = row_num % 2 == 0

        for col in range(1, COL_ROWNUM + 1):
            cell = ws.cell(row=row, column=col)

            if col == COL_ROWNUM:
                # Row-number stamp column (pre-filled, gray, locked visually)
                cell.value     = row_num
                cell.font      = _font(size=8, color="666666")
                cell.fill      = _fill(_GRAY)
                cell.alignment = _align("center", "center")
                cell.border    = Border(
                    left=_thin, right=_thick,
                    top=_thin, bottom=_thin,
                )
            else:
                cell.fill      = _fill(_ROW_A if even else _ROW_B)
                cell.alignment = _align("center", "center")
                cell.font      = _font(size=10)

                l = _thick if col == 1 else _thin
                r = _thin
                t = _thin
                b = _thin
                cell.border = Border(left=l, right=r, top=t, bottom=b)

    # Bottom edge on last data row
    for col in range(1, COL_ROWNUM + 1):
        cell = ws.cell(row=DATA_END, column=col)
        cell.border = Border(
            left=cell.border.left,
            right=cell.border.right,
            top=_thin,
            bottom=_thick,
        )


# ---------------------------------------------------------------------------
# Write bar data into an already-built sheet (called from excel_bridge)
# ---------------------------------------------------------------------------

def clear_data_area(ws) -> None:
    """Clear only the bar data rows (10-69), preserving the header."""
    for row in range(DATA_START, DATA_END + 1):
        for col in range(1, COL_ROWNUM):   # leave row-number column alone
            ws.cell(row=row, column=col).value = None


def bar_to_vista_row(bar) -> list:
    """
    Map one BarRow to an ordered list matching the Vista Steel columns A-Q
    (excludes the row-number column R, which is pre-filled).

    Returns a 17-element list: [SHL, UNITS, PER_UNIT, TOTAL, SIZE, GRADE,
                                 LENGTH, MARK, TYPE, A, B, C, D, E, F, G, H]
    """
    shl       = SHL_MAP.get(bar.shape, "S")
    bend_type = TYPE_MAP.get(bar.shape, "0")

    return [
        shl,               # A: S/H/L
        1,                 # B: NO. OF UNITS (1 assembly per mark)
        bar.qty,           # C: NO. PER UNIT
        bar.qty,           # D: TOTAL
        bar.size,          # E: SIZE
        "60",              # F: GRADE (ASTM A615/A706 Gr 60)
        bar.length_ft_in,  # G: LENGTH (cut length, formatted)
        bar.mark,          # H: MARK
        bend_type,         # I: TYPE
        bar.leg_a_ft_in or "",  # J: A
        bar.leg_b_ft_in or "",  # K: B
        bar.leg_c_ft_in or "",  # L: C'
        "",                # M: D (not computed for simple shapes)
        "",                # N: E
        "",                # O: F
        "",                # P: G
        "",                # Q: H
    ]

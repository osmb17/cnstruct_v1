"""
ReasoningLog tab — static openpyxl placeholder builder.

Live computation trace rows are written by ReasoningLogger via xlwings at runtime.
This module sets up the 5-column header and styling.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

_NAVY  = "1C3461"
_BLUE  = "2E75B6"
_WHITE = "FFFFFF"
_GRAY  = "F2F2F2"

_thin = Side(style="thin")
_med  = Side(style="medium")


def _fill(h: str) -> PatternFill:
    return PatternFill("solid", fgColor=h)


def _font(size=10, bold=False, color="000000", italic=False) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color, italic=italic)


def _align(h="left", v="center", wrap=False) -> Alignment:
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)


def _border(l=_thin, r=_thin, t=_thin, b=_thin) -> Border:
    return Border(left=l, right=r, top=t, bottom=b)


_COL_WIDTHS = {
    "A": 10,   # Timestamp
    "B": 10,   # Tag
    "C": 52,   # Message
    "D": 42,   # Detail / Formula
    "E": 16,   # Source
}

_COL_HEADERS = ["Timestamp", "Tag", "Message", "Detail / Formula", "Source"]


def build_reasoning_sheet(ws) -> None:
    """
    Rebuild *ws* (openpyxl Worksheet) as the ReasoningLog placeholder.
    """
    # Unmerge any existing merges
    for m in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(m))

    ws.sheet_view.showGridLines = False

    # Column widths
    for col, width in _COL_WIDTHS.items():
        ws.column_dimensions[col].width = width

    # ── Row 1: Title banner ──────────────────────────────────────────────
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = "AI REASONING LOG  |  Computation Trace"
    c.font      = _font(size=14, bold=True, color=_WHITE)
    c.fill      = _fill(_NAVY)
    c.alignment = _align("center", "center")

    # ── Row 2: Subtitle ──────────────────────────────────────────────────
    ws.row_dimensions[2].height = 18
    ws.merge_cells("A2:E2")
    c = ws["A2"]
    c.value     = ("Real-time step-by-step computation log. "
                   "Yellow rows = AI notes. Orange rows = warnings.")
    c.font      = _font(size=9, italic=True, color=_WHITE)
    c.fill      = _fill(_BLUE)
    c.alignment = _align("center", "center")

    # ── Row 3: spacer ────────────────────────────────────────────────────
    ws.row_dimensions[3].height = 8
    ws.merge_cells("A3:E3")

    # ── Row 4: Legend row ────────────────────────────────────────────────
    ws.row_dimensions[4].height = 18
    legend = [
        ("A4", "INIT",    "292929"),
        ("B4", "CALC",    "292929"),
        ("C4", "RULE",    "292929"),
        ("D4", "AI ✦",    "292929"),   # yellow
        ("E4", "WARN",    "292929"),   # orange
    ]
    legend_fills = {
        "A4": "F2F2F2", "B4": "F2F2F2", "C4": "F2F2F2",
        "D4": "FFFF99", "E4": "FFCC88",
    }
    for cell_ref, label, color in legend:
        c = ws[cell_ref]
        c.value     = f"● {label}"
        c.font      = _font(size=8, color="444444")
        c.fill      = _fill(legend_fills[cell_ref])
        c.alignment = _align("center", "center")
        c.border    = _border()

    # ── Row 5: spacer ────────────────────────────────────────────────────
    ws.row_dimensions[5].height = 6

    # ── Row 6: Column headers ────────────────────────────────────────────
    ws.row_dimensions[6].height = 22
    for i, label in enumerate(_COL_HEADERS):
        c = ws.cell(row=6, column=i + 1)
        c.value     = label
        c.font      = _font(size=10, bold=True, color=_WHITE)
        c.fill      = _fill(_NAVY)
        c.alignment = _align("center", "center")
        c.border    = _border(l=_med, r=_med, t=_med, b=_med)

    ws.freeze_panes = "A7"
